from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

def create_plan_pdf(user, diet_plan_html, workout_plan_html):
    """Generates a PDF of the simple AI-generated daily plans."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, f"Daily Plan for {user.name}", 0, 1, 'C')
    pdf.ln(10)

    pdf.set_font("Arial", 'B', 14)
    pdf.cell(0, 10, "Today's Diet Plan", 0, 1)
    pdf.set_font("Arial", '', 12)
    for item in diet_plan_html.split('</li>'):
        if 'plan-item' in item:
            try:
                name = item.split('<div class="item-name">')[1].split('</div>')[0]
                info = item.split('<div class="item-info">')[1].split('</div>')[0]
                pdf.cell(0, 8, f"- {name.strip()} ({info.strip()})", 0, 1)
            except IndexError:
                continue
    pdf.ln(10)

    pdf.set_font("Arial", 'B', 14)
    pdf.cell(0, 10, "Today's Workout Plan", 0, 1)
    pdf.set_font("Arial", '', 12)
    for item in workout_plan_html.split('</li>'):
        if 'plan-item' in item:
            try:
                name = item.split('<div class="item-name">')[1].split('</div>')[0]
                info = item.split('<div class="item-info">')[1].split('</div>')[0]
                pdf.cell(0, 8, f"- {name.strip()} ({info.strip()})", 0, 1)
            except IndexError:
                continue

    return pdf.output(dest='S').encode('latin-1')


def create_daily_plan_excel(diet_plan_html, workout_plan_html):
    """Generates an Excel file of the simple AI-generated daily plan."""
    workbook = openpyxl.Workbook()
    
    diet_sheet = workbook.active
    diet_sheet.title = "Today's Diet Plan"
    diet_headers = ["Meal", "To Be Eaten", "Calories"]
    diet_sheet.append(diet_headers)

    for item in diet_plan_html.split('</li>'):
        if 'plan-item' in item:
            try:
                full_name = item.split('<div class="item-name">')[1].split('</div>')[0].strip()
                calories = item.split('<div class="item-info">')[1].split(' kcal</div>')[0].strip()
                parts = full_name.split(': ', 1)
                meal_type = parts[0]
                food_name = parts[1] if len(parts) > 1 else ''
                diet_sheet.append([meal_type, food_name, int(calories)])
            except (IndexError, ValueError):
                continue

    workout_sheet = workbook.create_sheet("Today's Workout Plan")
    workout_headers = ["Category", "Exercise", "Calories Burned"]
    workout_sheet.append(workout_headers)

    for item in workout_plan_html.split('</li>'):
        if 'plan-item' in item:
            try:
                full_name = item.split('<div class="item-name">')[1].split('</div>')[0].strip()
                calories = item.split('<div class="item-info">')[1].split(' kcal burned</div>')[0].strip()
                parts = full_name.split(': ', 1)
                category = parts[0]
                exercise = parts[1] if len(parts) > 1 else ''
                workout_sheet.append([category, exercise, int(calories)])
            except (IndexError, ValueError):
                continue

    # --- THIS IS THE COMPLETE STYLING CODE ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="14B8A6", end_color="14B8A6", fill_type="solid")
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    return excel_file